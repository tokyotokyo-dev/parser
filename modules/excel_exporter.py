import logging
from typing import List, Dict

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl не установлен. Используем fallback через csv.")

import csv


COLUMNS = [
    ("Название канала",  "title",        30),
    ("Ссылка на канал",  "url",          45),
    ("Email",            "email",        35),
    ("Подписчики",       "subscribers",  15),
    ("Страна",           "country",      10),
    ("Видео",            "video_count",  10),
    ("Последний пост",   "last_upload",  18),
]

HEADER_COLOR = "1F3864"   # тёмно-синий
EMAIL_COLOR  = "E8F5E9"   # светло-зелёный
ALT_COLOR    = "F5F5F5"   # светло-серый


def _border():
    thin = Side(style="thin", color="BDBDBD")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def export_xlsx(channels: List[Dict], filepath: str) -> str:
    #Сохранение список каналов в .xlsx с форматированием.
    if not HAS_OPENPYXL:
        return _export_csv(channels, filepath.replace(".xlsx", ".csv"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "YouTube Channels"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (label, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # Данные
    channels = [ch for ch in channels if ch.get("email")]
    for row_idx, ch in enumerate(channels, start=2):
        is_alt = (row_idx % 2 == 0)
        has_email = bool(ch.get("email"))

        for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
            value = ch.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _border()
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            # Цвет строки
            if has_email:
                cell.fill = PatternFill("solid", fgColor=EMAIL_COLOR)
            elif is_alt:
                cell.fill = PatternFill("solid", fgColor=ALT_COLOR)

            # Ссылка — кликабельная
            if key == "url" and value:
                cell.hyperlink = value
                cell.font = Font(color="1565C0", underline="single")

            if key in ("subscribers", "video_count"):
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row_idx].height = 18

    # Итоговая строка
    total_row = len(channels) + 2
    ws.cell(row=total_row, column=1, value=f"Итого: {len(channels)} каналов").font = Font(bold=True)
    emails_found = sum(1 for ch in channels if ch.get("email"))
    ws.cell(row=total_row, column=3, value=f"С email: {emails_found}").font = Font(bold=True, color="2E7D32")

    # Фильтр
    ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)
    logger.info(f"[EXPORT] Сохранено в {filepath} | {len(channels)} строк | {emails_found} email")
    return filepath


def _export_csv(channels: List[Dict], filepath: str) -> str:
    #Сохранить в CSV, если openpyxl недоступен.
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f, fieldnames=[key for _, key, _ in COLUMNS]
        )
        writer.writeheader()
        writer.writerows(channels)
    logger.info(f"[EXPORT] CSV сохранён в {filepath}")
    return filepath
